import io
import os, sys
import re, json
import openpyxl
import multiprocessing, subprocess
import logging
from openpyxl.cell.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl import utils

fmt_str = "%(asctime)s: %(levelname)s: %(funcName)s Line:%(lineno)d %(message)s"
date_str = "%d/%m/%Y %H:%M:%S"
logging.basicConfig(level=logging.DEBUG, format=fmt_str, datefmt=date_str)
logger = logging.getLogger(__name__)
PICS_SCHEMA = None
PIXIT_SCHEMA = None

# check if file exist
def check_file(file):
    if getattr(sys, 'frozen', False):
      application_path = os.path.dirname(sys.executable)
    elif __file__:
      application_path = os.path.dirname(__file__)
    file_path = os.path.abspath(os.path.join(application_path, file))
    if (os.path.exists(file_path)):
        return file_path
    else:
        logger.critical(f"File {file_path} is not existed")
        return None

def get_enum_schema(param, schema_type):
    temp_schema = None
    if schema_type:
        global PICS_SCHEMA
        temp_schema = PICS_SCHEMA
    else:
        global PIXIT_SCHEMA
        temp_schema = PIXIT_SCHEMA
    index_val = param[1]
    if (param[0] in temp_schema["properties"]):
        if "$ref" in temp_schema["properties"][param[0]]:
            ref = temp_schema
            for node in temp_schema["properties"][param[0]]["$ref"].split("/"):
                if (node != "#"):
                    ref = ref[node]
            if "enum" in ref:
                index_val = ref["enum"].index(param[1])
            else:
                logger.warning(f"$ref {temp_schema['properties'][param[0]]['$ref']} does not have enum val")
        else:
            logger.warning(f"Param {param[0]} does not have $ref property")
    else:
        logger.warning(f"Param {param[0]} is not existed in {temp_schema['$id']}")
    return {param[0]:index_val}

def get_param(input_str:str, schema_type):
    input_str = re.sub(r"[\r\n]+", "", input_str).strip()
    result = {}
    for item in input_str.split(","):
        temp = item.split("=")
        if len(temp) > 1:
            param_name = temp[0].strip()
            param_val = temp[1].strip()
            if param_val.isnumeric():
                result.update({param_name: param_val})
            elif param_val == "true":
                result.update({param_name: True})
            elif param_val == "false":
                result.update({param_name: False})
            else:
                result.update(get_enum_schema((param_name, param_val), schema_type))
    return result

def get_testcase(ws:Worksheet, testcase_name):
    for row_cell in ws.iter_rows(ws.min_row, ws.max_row, utils.column_index_from_string("B"), utils.column_index_from_string("B")):
        if (str(row_cell[0].value).strip() == testcase_name):
            logger.debug(f"{testcase_name} at {row_cell[0].coordinate}")
            # load config from F,G
            config = {}
            config["pics"] = get_param(ws.cell(row_cell[0].row, utils.column_index_from_string("F")).value, True) # true = pics
            config["pixit"] = get_param(ws.cell(row_cell[0].row, utils.column_index_from_string("G")).value, False) # false = pixit
            return [(testcase_name, config)]
    return []

def get_all_testcase(ws:Worksheet):
    testcase_list = []
    for select_cell in ws.iter_rows(ws.min_row, ws.max_row, utils.column_index_from_string("J"), utils.column_index_from_string("J")):
        if str(select_cell[0].value).strip() == "none":
            testcase_name = str(ws.cell(select_cell[0].row, utils.column_index_from_string("B")).value).strip()
            logger.debug(f"select tc: {testcase_name} at {select_cell[0].coordinate}")
            config = {}
            config["pics"] = get_param(ws.cell(select_cell[0].row, utils.column_index_from_string("F")).value, True) # true = pics
            config["pixit"] = get_param(ws.cell(select_cell[0].row, utils.column_index_from_string("G")).value, False) # false = pixit
            testcase_list.append((testcase_name, config))
    return testcase_list

# find testcase
def get_excel_tc(file, testcase_name=None):
    # process all enable testcase in excel file
    logger.info(f"Processing {file} ...")
    wb = openpyxl.load_workbook(file, read_only=True)
    ws = wb["testcase"]
    # get all enable testcase
    if (testcase_name == None):
        result = get_all_testcase(ws)
    else:
        result = get_testcase(ws, testcase_name)
    wb.close()
    return result

def update_testconfig(config, update):
    for _type in ["pics", "pixit"]:
        for k,v in update[_type].items():
            config[_type][k] = v

RETURN_CODE = {
    0:"none",
    1:"pass",
    2:"inconclude",
    3:"fail",
    4:"error"
}

def export_report(result, report_file):
    wb = openpyxl.load_workbook(report_file)
    ws = wb["testcase"]
    for row_cell in ws.iter_rows(ws.min_row, ws.max_row, utils.column_index_from_string("B"), utils.column_index_from_string("B")):
        tcname = str(row_cell[0].value).strip()
        if (tcname in result):
            ws.cell(row_cell[0].row, utils.column_index_from_string("J")).value = result[tcname]["result"]
            ws.cell(row_cell[0].row, utils.column_index_from_string("K")).value = result[tcname]["log"]
            logger.info(f"Testcase {tcname} - result: {result[tcname]['result']}")
    wb.save(os.path.join(os.path.dirname(report_file), "testcase.xlsx"))
    wb.close()

def main(argv):
    # build
    build_p = subprocess.Popen(["make", "-j", "2"], cwd=os.path.abspath("./build"))

    testcase_list = []
    excel_file = check_file("./testcase.xlsx")
    global PICS_SCHEMA
    global PIXIT_SCHEMA
    PICS_SCHEMA = json.load(open(check_file("./docs/pics_schema.json"), "r", encoding="utf8"))
    PIXIT_SCHEMA = json.load(open(check_file("./docs/pixit_schema.json"), "r", encoding="utf8"))
    if (excel_file == None):
        return -1
    if len(argv) == 1:
        # execute all testcase in testcase.xlsx, get all testcase
        testcase_list = get_excel_tc(excel_file)
    else:
        # execute specific testcase as input
        testcase_list = get_excel_tc(excel_file, argv[1])

    if (0 != build_p.wait()):
        logger.critical("Failed to build runTC executable")
        return -1

    testreport = {}
    # run testcase
    for tc, cfg in testcase_list:
        # generate config for testcase
        output_cfg_path = "./test.json"
        with open(check_file("./docs/testconfig.json"), "r", encoding="utf8") as fp:
            test_config = json.load(fp)
            # export testcase config dict to test_config.json file
            update_testconfig(test_config, cfg)
            with open(output_cfg_path, "w", encoding="utf8") as wfp:
                json.dump(test_config, wfp, indent=4)
        executable = check_file("./build/TestSuite/runTC")
        instance = subprocess.run([executable, tc, "I", output_cfg_path], stdout=subprocess.PIPE, universal_newlines=True)
        # ----- TEST CASE END -----
        # [MTC]: Verdict:
        #instance.stdout.seek(3000, os.SEEK_END)
        #get_flag = False
        log = instance.stdout.partition("TEST CASE END")[2]
        #for line in iter(instance.stdout.readline, ''):
        #    if get_flag:
        #        log += line
        #    if "TEST CASE END" in line:
        #        get_flag = True
        # parse result and log to store
        testreport.update({tc: {
            # none = 0, pass = 1, inconc = 2, fail = 3, error = 4
            "result":RETURN_CODE[instance.returncode],
            "log":log
        }})

    # update result to testcase_report.xlsx
    export_report(testreport, excel_file)
    return 0

if __name__ == "__main__":
    exit(main(sys.argv))
